#!/usr/bin/env python3
"""
cluster_blosum.py — Step 2a of the VHH NGS pipeline.

BLOSUM62-based hierarchical clustering of CDR3 (or full VHH) sequences.
Key improvements over original:

  • Switched from deprecated Bio.pairwise2 → Bio.Align.PairwiseAligner
  • Parallelised distance matrix computation (multiprocessing)
  • Progress bars for long-running distance calculations
  • Minimum cluster count filter (mirrors HDBSCAN output format)
  • Shannon entropy per cluster (diversity metric)
  • Weighted consensus (by read count) instead of majority vote
  • Excel output with Cluster_Count column for enrichment compatibility
  • Deterministic cluster IDs (sorted by total cluster count, descending)
"""

import argparse
import sys
import os
import csv
import logging
from pathlib import Path
from collections import Counter
from multiprocessing import Pool, cpu_count
from functools import partial
from typing import List, Tuple, Optional

import numpy as np
import pandas as pd
from scipy.cluster.hierarchy import linkage, fcluster
from Bio.Align import PairwiseAligner, substitution_matrices
from tqdm import tqdm
from rich.console import Console
from rich.table import Table
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

console = Console()

# ══════════════════════════════════════════════════════════════════════════════
# ALIGNER SETUP  (module-level so it can be pickled for multiprocessing)
# ══════════════════════════════════════════════════════════════════════════════

def _make_aligner() -> PairwiseAligner:
    aln = PairwiseAligner()
    aln.substitution_matrix = substitution_matrices.load("BLOSUM62")
    aln.open_gap_score   = -10
    aln.extend_gap_score = -0.5
    aln.mode = "global"
    return aln


def _blosum_distance(pair: Tuple[int, int], seqs: List[str]) -> Tuple[int, int, float]:
    """
    Worker: compute BLOSUM62 percent-identity distance for a sequence pair.
    Returns (i, j, distance) where distance = 100 - %identity.
    """
    i, j = pair
    aln = _make_aligner()
    alignments = aln.align(seqs[i], seqs[j])
    if not alignments:
        return i, j, 100.0
    best = next(iter(alignments))
    aligned_a, aligned_b = best[0], best[1]
    matches   = sum(a == b for a, b in zip(aligned_a, aligned_b) if a != "-" and b != "-")
    ungapped  = sum(1 for a, b in zip(aligned_a, aligned_b) if a != "-" and b != "-")
    identity  = (matches / ungapped * 100) if ungapped > 0 else 0
    return i, j, max(0.0, 100.0 - identity)


# ══════════════════════════════════════════════════════════════════════════════
# DISTANCE MATRIX
# ══════════════════════════════════════════════════════════════════════════════

def compute_distance_matrix(seqs: List[str], n_jobs: int = -1) -> np.ndarray:
    """
    Build pairwise BLOSUM62 distance matrix using parallel workers.
    n_jobs=-1 → all available CPUs.
    """
    n = len(seqs)
    pairs = [(i, j) for i in range(n) for j in range(i + 1, n)]
    n_cores = cpu_count() if n_jobs == -1 else max(1, n_jobs)

    dist = np.zeros((n, n))
    worker = partial(_blosum_distance, seqs=seqs)

    console.print(f"  Computing {len(pairs):,} pairwise distances "
                  f"on {n_cores} CPU cores...")
    with Pool(n_cores) as pool:
        for i, j, d in tqdm(pool.imap_unordered(worker, pairs, chunksize=50),
                             total=len(pairs), desc="Distance matrix", unit="pair"):
            dist[i, j] = d
            dist[j, i] = d
    return dist


# ══════════════════════════════════════════════════════════════════════════════
# CONSENSUS AND DIVERSITY
# ══════════════════════════════════════════════════════════════════════════════

def weighted_consensus(seqs: List[str], weights: Optional[List[int]] = None) -> str:
    """
    Build consensus by weighted majority vote per position.
    If weights not provided, all sequences weighted equally.
    """
    if not seqs:
        return ""
    if weights is None:
        weights = [1] * len(seqs)
    max_len = max(len(s) for s in seqs)
    padded  = [s.ljust(max_len, "-") for s in seqs]
    consensus = []
    for pos in range(max_len):
        vote: Counter = Counter()
        for seq, w in zip(padded, weights):
            vote[seq[pos]] += w
        consensus.append(vote.most_common(1)[0][0])
    return "".join(c for c in consensus if c != "-")


def shannon_entropy(seqs: List[str]) -> float:
    """Sequence-level Shannon entropy: H = -Σ p(s) log2 p(s)."""
    if len(seqs) <= 1:
        return 0.0
    total = len(seqs)
    counts = Counter(seqs)
    return -sum((c / total) * np.log2(c / total) for c in counts.values())


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

def write_excel_with_heatmap(df: pd.DataFrame, filepath: Path):
    df.to_excel(filepath, index=False)
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    last_row = ws.max_row
    if "Cluster_Count" in df.columns:
        col_idx    = list(df.columns).index("Cluster_Count") + 1
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{last_row}",
            ColorScaleRule(
                start_type="min",  start_color="FFFFFF",
                mid_type="percentile", mid_value=50, mid_color="FFFF99",
                end_type="max",    end_color="FF0000",
            ),
        )
    wb.save(filepath)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN CLUSTERING LOGIC
# ══════════════════════════════════════════════════════════════════════════════

def run_blosum_clustering(
    input_csv:       str,
    threshold:       float,
    seq_column:      str  = "CDR3",
    min_cluster_count: int = 5,
    n_jobs:          int  = -1,
    linkage_method:  str  = "average",   # UPGMA — appropriate for sequence clustering
):
    input_path = Path(input_csv)
    base       = input_path.with_suffix("")

    df = pd.read_csv(input_csv)
    if seq_column not in df.columns:
        console.print(f"[red]ERROR: column '{seq_column}' not found in CSV.[/red]")
        sys.exit(1)

    # ── Cleaning ───────────────────────────────────────────────────────────────
    df[seq_column] = df[seq_column].astype(str).str.upper().str.strip()
    before = len(df)
    df = df[
        (df[seq_column].str.len() > 0) &
        (~df[seq_column].str.contains("X")) &
        (~df[seq_column].str.contains(r"\*"))
    ].reset_index(drop=True)
    console.print(f"  Filtered {before - len(df)} invalid sequences → {len(df):,} remaining")

    if len(df) < 2:
        console.print("[yellow]Not enough sequences to cluster.[/yellow]")
        return

    sequences = df[seq_column].tolist()
    counts    = df["Count"].tolist() if "Count" in df.columns else [1] * len(df)

    # ── Distance matrix ────────────────────────────────────────────────────────
    dist = compute_distance_matrix(sequences, n_jobs=n_jobs)

    # ── Hierarchical clustering ────────────────────────────────────────────────
    console.print(f"  Clustering with '{linkage_method}' linkage, threshold = {threshold}%")
    Z        = linkage(dist, method=linkage_method)
    clusters = fcluster(Z, t=threshold, criterion="distance")
    df["Cluster"] = clusters

    # ── Apply minimum cluster count filter ────────────────────────────────────
    cluster_total_counts = df.groupby("Cluster")["Count"].sum() \
        if "Count" in df.columns \
        else df.groupby("Cluster").size()
    valid = cluster_total_counts[cluster_total_counts >= min_cluster_count].index
    df = df[df["Cluster"].isin(valid)].reset_index(drop=True)
    console.print(f"  {len(valid)} clusters pass min_cluster_count={min_cluster_count}")

    # ── Renumber clusters by total count (1 = largest) ────────────────────────
    rank_map = {
        old: new + 1
        for new, old in enumerate(
            cluster_total_counts[valid].sort_values(ascending=False).index
        )
    }
    df["Cluster"] = df["Cluster"].map(rank_map)

    # ── Per-member CSV ─────────────────────────────────────────────────────────
    cluster_csv = str(base) + "_blosum_clusters.csv"
    df.to_csv(cluster_csv, index=False)

    # ── Consensus CSV & Excel ──────────────────────────────────────────────────
    consensus_rows = []
    for cid in sorted(df["Cluster"].unique()):
        sub      = df[df["Cluster"] == cid]
        seqs_c   = sub[seq_column].tolist()
        wts_c    = sub["Count"].tolist() if "Count" in sub.columns else None
        consensus = weighted_consensus(seqs_c, wts_c)
        entropy   = shannon_entropy(seqs_c)
        total_cnt = sub["Count"].sum() if "Count" in sub.columns else len(sub)
        n_unique  = sub[seq_column].nunique()
        consensus_rows.append({
            "Cluster":          cid,
            "CDR3":             consensus,          # canonical field for enrichment step
            "Cluster_Count":    int(total_cnt),     # required by Cluster_enrichment.py
            "Members":          len(sub),
            "Unique_Seqs":      n_unique,
            "Shannon_Entropy":  round(entropy, 3),
        })

    consensus_df = pd.DataFrame(consensus_rows)
    consensus_csv  = str(base) + "_cluster_consensus.csv"
    consensus_xlsx = str(base) + "_cluster_consensus.xlsx"

    consensus_df.to_csv(consensus_csv, index=False)
    write_excel_with_heatmap(consensus_df, Path(consensus_xlsx))

    # ── Summary ────────────────────────────────────────────────────────────────
    table = Table(title="Clustering Summary", show_lines=True)
    table.add_column("Metric", style="cyan")
    table.add_column("Value",  justify="right")
    table.add_row("Input sequences",    str(len(sequences)))
    table.add_row("Sequences clustered", str(len(df)))
    table.add_row("Clusters",           str(len(consensus_rows)))
    table.add_row("Threshold",          f"{threshold}% distance")
    table.add_row("Linkage method",     linkage_method)
    console.print(table)
    console.print(f"\n[green]✓[/green] {cluster_csv}")
    console.print(f"[green]✓[/green] {consensus_csv}")
    console.print(f"[green]✓[/green] {consensus_xlsx}")


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="BLOSUM62 hierarchical clustering for VHH CDR3 sequences"
    )
    parser.add_argument("--input",    required=True,  help="Input CSV with sequences")
    parser.add_argument("--column",   default="CDR3",  help="Sequence column (default: CDR3)")
    parser.add_argument("--threshold", type=float, required=True,
                        help="Distance threshold for cluster cutoff (0–100; lower = stricter)")
    parser.add_argument("--min-count", type=int, default=5,
                        help="Min total read count to report a cluster (default: 5)")
    parser.add_argument("--jobs",      type=int, default=-1,
                        help="CPU cores for distance matrix (-1 = all, default: -1)")
    parser.add_argument("--linkage",   default="average",
                        choices=["single", "complete", "average", "ward"],
                        help="Linkage method (default: average/UPGMA)")
    args = parser.parse_args()

    if args.threshold is None:
        console.print("[red]ERROR: --threshold is required[/red]")
        sys.exit(1)

    run_blosum_clustering(
        args.input, args.threshold, args.column,
        min_cluster_count=args.min_count,
        n_jobs=args.jobs,
        linkage_method=args.linkage,
    )


if __name__ == "__main__":
    main()
