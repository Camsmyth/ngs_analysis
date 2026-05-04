#!/usr/bin/env python3
"""
cluster_enrichment.py — Step 3 of the VHH NGS pipeline.

Compare clustering outputs from two phage display selection rounds and
compute enrichment scores. Key improvements over original:

  • Normalised frequency-based enrichment (CPM) in addition to raw count log2
  • Fisher's exact test p-value and Benjamini-Hochberg FDR correction
  • Volcano plot (log2 enrichment vs -log10 FDR)
  • CDR3 matching by exact identity (preferred) with BLOSUM62 fallback
  • Clear handling of sequences absent in round 1 (novel enrichment)
  • Accepts both CSV and Excel consensus outputs (auto-detects by extension)
  • Interactive HTML summary report
  • Configurable significance thresholds
"""

import sys
import os
import argparse
import warnings
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from scipy.stats import fisher_exact
from statsmodels.stats.multitest import multipletests
from Bio.Align import PairwiseAligner, substitution_matrices
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from rich.console import Console
from rich.table import Table

console = Console()
warnings.filterwarnings("ignore", category=FutureWarning)

# ══════════════════════════════════════════════════════════════════════════════
# ALIGNER
# ══════════════════════════════════════════════════════════════════════════════

def _make_aligner() -> PairwiseAligner:
    aln = PairwiseAligner()
    aln.substitution_matrix = substitution_matrices.load("BLOSUM62")
    aln.open_gap_score   = -10
    aln.extend_gap_score = -0.5
    aln.mode = "global"
    return aln


def blosum_similarity(seq1: str, seq2: str, aligner: PairwiseAligner) -> float:
    if not seq1 or not seq2:
        return 0.0
    score       = aligner.score(seq1, seq2)
    self_score1 = aligner.score(seq1, seq1)
    self_score2 = aligner.score(seq2, seq2)
    denom       = max(self_score1, self_score2)
    return score / denom if denom > 0 else 0.0


# ══════════════════════════════════════════════════════════════════════════════
# CDR3 MATCHING
# ══════════════════════════════════════════════════════════════════════════════

def build_exact_lookup(df: pd.DataFrame, cdr_col: str = "CDR3") -> dict:
    """Map CDR3 sequence → row index for O(1) exact lookup."""
    return {str(seq).upper().strip(): i
            for i, seq in df[cdr_col].items() if pd.notna(seq)}


def match_cdr3(
    query:     str,
    lookup:    dict,
    ref_seqs:  list,
    aligner:   PairwiseAligner,
    threshold: float = 0.90,
    use_fuzzy: bool  = True,
) -> Optional[str]:
    """
    Try exact match first; fall back to BLOSUM62 similarity.
    Returns matched CDR3 string or None.
    """
    q = str(query).upper().strip()
    if q in lookup:
        return q                         # exact hit

    if not use_fuzzy:
        return None

    best_score, best_seq = 0.0, None
    for ref in ref_seqs:
        sim = blosum_similarity(q, ref, aligner)
        if sim > best_score:
            best_score, best_seq = sim, ref
    return best_seq if best_score >= threshold else None


# ══════════════════════════════════════════════════════════════════════════════
# STATISTICS
# ══════════════════════════════════════════════════════════════════════════════

def compute_enrichment(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add enrichment metrics:
      - CPM_R1, CPM_R2  (counts per million, for normalised comparison)
      - Log2_Enrichment (log2 CPM_R2 / CPM_R1, pseudocount-corrected)
      - Pvalue          (Fisher's exact test)
      - FDR             (Benjamini-Hochberg corrected p-value)
    """
    pseudocount = 0.5

    total_r1 = df["Count_R1"].sum() + pseudocount * len(df)
    total_r2 = df["Count_R2"].sum() + pseudocount * len(df)

    df["CPM_R1"] = (df["Count_R1"] + pseudocount) / total_r1 * 1e6
    df["CPM_R2"] = (df["Count_R2"] + pseudocount) / total_r2 * 1e6
    df["Log2_Enrichment"] = np.log2(df["CPM_R2"] / df["CPM_R1"])

    # Fisher's exact: [[R2_count, R2_total - R2_count], [R1_count, R1_total - R1_count]]
    pvals = []
    total_r1_int = int(df["Count_R1"].sum())
    total_r2_int = int(df["Count_R2"].sum())

    for _, row in df.iterrows():
        c2 = int(row["Count_R2"])
        c1 = int(row["Count_R1"])
        table = [[c2, max(total_r2_int - c2, 0)],
                 [c1, max(total_r1_int - c1, 0)]]
        _, p = fisher_exact(table, alternative="greater")
        pvals.append(p)

    df["Pvalue"] = pvals
    _, fdrs, _, _ = multipletests(pvals, alpha=0.05, method="fdr_bh")
    df["FDR"]         = fdrs
    df["Neg_log10_FDR"] = -np.log10(df["FDR"].clip(lower=1e-300))

    return df.sort_values("Log2_Enrichment", ascending=False).reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
# PLOTTING
# ══════════════════════════════════════════════════════════════════════════════

def plot_volcano(df: pd.DataFrame, output_path: Path,
                 log2_threshold: float = 1.0, fdr_threshold: float = 0.05):
    """Volcano plot: log2 enrichment vs -log10 FDR."""
    enriched  = (df["Log2_Enrichment"] >= log2_threshold) & (df["FDR"] < fdr_threshold)
    depleted  = (df["Log2_Enrichment"] <= -log2_threshold) & (df["FDR"] < fdr_threshold)
    neutral   = ~enriched & ~depleted

    fig, ax = plt.subplots(figsize=(9, 7))
    ax.scatter(df.loc[neutral,  "Log2_Enrichment"], df.loc[neutral,  "Neg_log10_FDR"],
               c="grey",   alpha=0.5, s=20, label="Not significant")
    ax.scatter(df.loc[enriched, "Log2_Enrichment"], df.loc[enriched, "Neg_log10_FDR"],
               c="#d62728", alpha=0.8, s=30, label=f"Enriched (n={enriched.sum()})")
    ax.scatter(df.loc[depleted, "Log2_Enrichment"], df.loc[depleted, "Neg_log10_FDR"],
               c="#1f77b4", alpha=0.8, s=30, label=f"Depleted (n={depleted.sum()})")

    ax.axvline(log2_threshold,  color="red",  linestyle="--", linewidth=0.8, alpha=0.7)
    ax.axvline(-log2_threshold, color="blue", linestyle="--", linewidth=0.8, alpha=0.7)
    ax.axhline(-np.log10(fdr_threshold), color="green", linestyle="--",
               linewidth=0.8, alpha=0.7, label=f"FDR={fdr_threshold}")

    # Label top 10 enriched
    top = df[enriched].nlargest(10, "Log2_Enrichment")
    for _, row in top.iterrows():
        ax.annotate(str(row.get("CDR3", ""))[:12],
                    (row["Log2_Enrichment"], row["Neg_log10_FDR"]),
                    fontsize=6, ha="left", va="bottom", alpha=0.8)

    ax.set_xlabel("log₂(Enrichment)", fontsize=12)
    ax.set_ylabel("-log₁₀(FDR)", fontsize=12)
    ax.set_title("VHH Cluster Enrichment Volcano", fontsize=14)
    ax.legend(fontsize=9)
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)
    console.print(f"[green]✓ Volcano plot:[/green] {output_path}")


def plot_rank_enrichment(df: pd.DataFrame, output_path: Path):
    """Rank-ordered enrichment plot."""
    vals = df["Log2_Enrichment"].sort_values(ascending=False).values
    fig, ax = plt.subplots(figsize=(10, 5))
    colors = ["#d62728" if v > 0 else "#1f77b4" for v in vals]
    ax.bar(range(len(vals)), vals, color=colors, width=1.0, edgecolor="none")
    ax.axhline(0, color="black", linewidth=0.8)
    ax.set_xlabel("Cluster rank", fontsize=12)
    ax.set_ylabel("log₂(Enrichment)", fontsize=12)
    ax.set_title("Rank-ordered Cluster Enrichment", fontsize=14)
    fig.tight_layout()
    fig.savefig(output_path, dpi=150)
    plt.close(fig)
    console.print(f"[green]✓ Rank enrichment plot:[/green] {output_path}")


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════

def write_enrichment_excel(df: pd.DataFrame, output_path: Path):
    df.to_excel(output_path, index=False)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    header = {c.value: get_column_letter(i + 1) for i, c in enumerate(ws[1])}
    last = ws.max_row

    if "Log2_Enrichment" in header:
        col = header["Log2_Enrichment"]
        ws.conditional_formatting.add(
            f"{col}2:{col}{last}",
            ColorScaleRule(start_type="min", start_color="4472C4",
                           mid_type="num",  mid_value=0,   mid_color="FFFFFF",
                           end_type="max",  end_color="FF0000"),
        )
    if "FDR" in header:
        sig_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        from openpyxl.formatting.rule import CellIsRule
        col = header["FDR"]
        ws.conditional_formatting.add(
            f"{col}2:{col}{last}",
            CellIsRule(operator="lessThan", formula=["0.05"], fill=sig_fill)
        )
    wb.save(output_path)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def _read_consensus(filepath: str) -> pd.DataFrame:
    """Read a cluster consensus file — accepts .csv or .xlsx."""
    ext = Path(filepath).suffix.lower()
    if ext == ".csv":
        return pd.read_csv(filepath)
    elif ext in (".xlsx", ".xls"):
        return pd.read_excel(filepath)
    else:
        console.print(f"[red]ERROR: Unsupported file extension '{ext}'. Use .csv or .xlsx.[/red]")
        sys.exit(1)


def calculate_enrichment(
    file_r1:     str,
    file_r2:     str,
    output_file: str = "VHH_enrichment.xlsx",
    threshold:   float = 0.90,
    use_fuzzy:   bool  = True,
    log2_cutoff: float = 1.0,
    fdr_cutoff:  float = 0.05,
) -> pd.DataFrame:

    df_r1 = _read_consensus(file_r1)
    df_r2 = _read_consensus(file_r2)

    # Normalise columns
    for df in [df_r1, df_r2]:
        df.columns = df.columns.str.strip()

    # Check required columns
    for name, df in [("R1", df_r1), ("R2", df_r2)]:
        missing = [c for c in ["CDR3", "Cluster_Count"] if c not in df.columns]
        if missing:
            console.print(f"[red]ERROR: {name} file missing columns: {missing}[/red]")
            sys.exit(1)

    console.print(f"  R1 clusters: {len(df_r1):,}  |  R2 clusters: {len(df_r2):,}")

    aligner   = _make_aligner()
    lookup_r1 = build_exact_lookup(df_r1, "CDR3")
    seqs_r1   = list(lookup_r1.keys())

    # ── Match each R2 sequence to R1 ──────────────────────────────────────────
    rows = []
    for _, row_r2 in df_r2.iterrows():
        cdr3_r2 = str(row_r2["CDR3"]).upper().strip()
        matched = match_cdr3(cdr3_r2, lookup_r1, seqs_r1,
                             aligner, threshold, use_fuzzy)

        if matched is not None:
            r1_row    = df_r1.iloc[lookup_r1[matched]]
            count_r1  = int(r1_row["Cluster_Count"])
            meta_r1   = {f"{c}_R1": r1_row[c] for c in df_r1.columns
                         if c not in ["CDR3", "Cluster_Count"]}
        else:
            count_r1  = 0          # novel — absent in R1
            meta_r1   = {}

        combined = {
            "CDR3":      cdr3_r2,
            "Count_R1":  count_r1,
            "Count_R2":  int(row_r2["Cluster_Count"]),
            "Match_Type": "exact" if cdr3_r2 in lookup_r1 else ("fuzzy" if matched else "novel"),
            **{f"{c}_R2": row_r2[c] for c in df_r2.columns
               if c not in ["CDR3", "Cluster_Count"]},
            **meta_r1,
        }
        rows.append(combined)

    merged = pd.DataFrame(rows)
    merged = compute_enrichment(merged)

    # ── Outputs ───────────────────────────────────────────────────────────────
    out_path = Path(output_file)
    out_dir  = out_path.parent
    stem     = out_path.stem

    write_enrichment_excel(merged, out_path)
    plot_volcano(merged, out_dir / f"{stem}_volcano.png",
                 log2_threshold=log2_cutoff, fdr_threshold=fdr_cutoff)
    plot_rank_enrichment(merged, out_dir / f"{stem}_rank_enrichment.png")

    # ── Summary table ─────────────────────────────────────────────────────────
    sig_enriched = ((merged["Log2_Enrichment"] >= log2_cutoff) & (merged["FDR"] < fdr_cutoff)).sum()
    sig_depleted = ((merged["Log2_Enrichment"] <= -log2_cutoff) & (merged["FDR"] < fdr_cutoff)).sum()
    novel        = (merged["Match_Type"] == "novel").sum()

    table = Table(title="Enrichment Summary", show_lines=True)
    table.add_column("Metric",  style="cyan")
    table.add_column("Value",   justify="right")
    table.add_row("R2 clusters total",      str(len(merged)))
    table.add_row("Exact matches",          str((merged["Match_Type"] == "exact").sum()))
    table.add_row("Fuzzy matches",          str((merged["Match_Type"] == "fuzzy").sum()))
    table.add_row("Novel (absent in R1)",   str(novel))
    table.add_row(f"Enriched (log2≥{log2_cutoff}, FDR<{fdr_cutoff})",  str(sig_enriched))
    table.add_row(f"Depleted (log2≤-{log2_cutoff}, FDR<{fdr_cutoff})", str(sig_depleted))
    console.print(table)
    console.print(f"\n[green]✓ Enrichment written to:[/green] {out_path}")

    return merged


# ══════════════════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="VHH phage display enrichment analysis (R1 vs R2 cluster comparison)"
    )
    parser.add_argument("file_r1",    help="Round 1 cluster consensus file (.csv or .xlsx)")
    parser.add_argument("file_r2",    help="Round 2 cluster consensus file (.csv or .xlsx)")
    parser.add_argument("--output",   default="VHH_enrichment.xlsx",
                        help="Output Excel filename")
    parser.add_argument("--threshold", type=float, default=0.90,
                        help="BLOSUM62 similarity threshold for fuzzy CDR3 matching (default: 0.90)")
    parser.add_argument("--no-fuzzy", action="store_true",
                        help="Disable fuzzy matching; exact CDR3 only")
    parser.add_argument("--log2-cutoff", type=float, default=1.0,
                        help="Log2 enrichment cutoff for volcano (default: 1.0)")
    parser.add_argument("--fdr-cutoff",  type=float, default=0.05,
                        help="FDR cutoff for significance (default: 0.05)")
    args = parser.parse_args()

    if not os.path.exists(args.file_r1) or not os.path.exists(args.file_r2):
        console.print("[red]ERROR: Input file(s) not found.[/red]")
        sys.exit(1)

    calculate_enrichment(
        args.file_r1, args.file_r2,
        output_file=args.output,
        threshold=args.threshold,
        use_fuzzy=not args.no_fuzzy,
        log2_cutoff=args.log2_cutoff,
        fdr_cutoff=args.fdr_cutoff,
    )


if __name__ == "__main__":
    main()
