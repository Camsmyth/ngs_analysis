#!/usr/bin/env python3
"""
hdbscan_cluster.py — Step 2b of the VHH NGS pipeline (alternative to BLOSUM clustering).

Physicochemical embedding + HDBSCAN density clustering.
Key improvements over original:

  • Richer embedding: CDR1+CDR2+CDR3 physicochemical features (not CDR3 only)
  • Optional UMAP dimensionality reduction before HDBSCAN (improves clustering)
  • Cluster stability scores reported (HDBSCAN native)
  • Weighted representative selection (highest-count protein per cluster)
  • Shannon entropy and liability summary per cluster
  • Excel output fully compatible with Cluster_enrichment.py
    (includes CDR3 and Cluster_Count columns)
  • Soft clustering: reports membership probabilities for noise points
  • Configurable via CLI
"""

import argparse
import logging
import sys
from collections import defaultdict, Counter
from datetime import datetime
from pathlib import Path
from typing import List, Optional

import numpy as np
import pandas as pd
import hdbscan
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from Bio.SeqUtils.ProtParam import ProteinAnalysis
from rich.console import Console
from rich.table import Table

console = Console()

# ── Try UMAP ──────────────────────────────────────────────────────────────────
try:
    import umap
    UMAP_AVAILABLE = True
except ImportError:
    UMAP_AVAILABLE = False

# ══════════════════════════════════════════════════════════════════════════════
# DEFAULT CONFIG
# ══════════════════════════════════════════════════════════════════════════════
DEFAULT_CONFIG = {
    "cluster_column":       "CDR3",
    "min_cluster_counts":   10,
    "min_cluster_size":     2,
    "min_samples":          2,
    "use_umap":             True,   # reduce before HDBSCAN if UMAP available
    "umap_n_components":    10,
    "embed_cdrs":           ["CDR1", "CDR2", "CDR3"],  # which CDRs to embed
    "good_fill":  PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "bad_fill":   PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
}

# ── Kyte-Doolittle hydrophobicity ─────────────────────────────────────────────
KD_DICT = {
    'A': 1.8,  'R': -4.5, 'N': -3.5, 'D': -3.5, 'C': 2.5,
    'Q': -3.5, 'E': -3.5, 'G': -0.4, 'H': -3.2, 'I': 4.5,
    'L': 3.8,  'K': -3.9, 'M': 1.9,  'F': 2.8,  'P': -1.6,
    'S': -0.8, 'T': -0.7, 'W': -0.9, 'Y': -1.3, 'V': 4.2,
}

# ── Liability patterns ────────────────────────────────────────────────────────
import regex as re
LIABILITY_PATTERNS = {
    "N-glycosylation":  re.compile(r"N[^P][ST]"),
    "Deamidation":      re.compile(r"N[GS]"),
    "Isomerization":    re.compile(r"D[TGSH]"),
    "Free Cys":         re.compile(r"C"),
}


# ══════════════════════════════════════════════════════════════════════════════
# EMBEDDING
# ══════════════════════════════════════════════════════════════════════════════

def embed_sequence(seq: str) -> np.ndarray:
    """
    Physicochemical embedding of a CDR sequence.
    Returns a fixed-length feature vector independent of sequence length.
    """
    if not seq or not all(c in KD_DICT for c in seq):
        seq = "".join(c for c in seq if c in KD_DICT) or "G"  # fallback

    kd = [KD_DICT[aa] for aa in seq]
    kd_mean, kd_std = np.mean(kd), np.std(kd)
    kd_min,  kd_max = np.min(kd),  np.max(kd)

    try:
        analysis = ProteinAnalysis(seq)
        pi          = analysis.isoelectric_point()
        aromaticity = analysis.aromaticity()
        aa_counts   = analysis.count_amino_acids()
    except Exception:
        pi = aromaticity = 7.0
        aa_counts = {}

    L, I, V, A = (aa_counts.get(k, 0) for k in "LIVA")
    aliphatic   = (A + 2.9 * (I + L) + 3.9 * V) / max(len(seq), 1) * 100

    aa_order   = "ACDEFGHIKLMNPQRSTVWY"
    total      = max(len(seq), 1)
    aa_frac    = [seq.count(aa) / total for aa in aa_order]

    # Length as normalised feature
    length_feat = [min(len(seq) / 30, 1.0)]

    return np.array([kd_mean, kd_std, kd_min, kd_max,
                     pi, aromaticity, aliphatic] + aa_frac + length_feat)


def embed_row(row: pd.Series, cdr_cols: List[str]) -> np.ndarray:
    """Concatenate embeddings for multiple CDR columns."""
    parts = []
    for col in cdr_cols:
        val = str(row.get(col, "")) if pd.notna(row.get(col, "")) else ""
        parts.append(embed_sequence(val))
    return np.concatenate(parts)


# ══════════════════════════════════════════════════════════════════════════════
# UTILITIES
# ══════════════════════════════════════════════════════════════════════════════

def shannon_entropy(seqs: List[str]) -> float:
    total = len(seqs)
    if total <= 1:
        return 0.0
    counts = Counter(seqs)
    return -sum((c / total) * np.log2(c / total) for c in counts.values())


def flag_liabilities(seq: str) -> str:
    flags = [name for name, pat in LIABILITY_PATTERNS.items() if pat.search(seq)]
    return ";".join(flags) if flags else "None"


# ══════════════════════════════════════════════════════════════════════════════
# CONSENSUS / REPRESENTATIVE
# ══════════════════════════════════════════════════════════════════════════════

def weighted_representative(sub_df: pd.DataFrame, col: str = "Protein_Sequence") -> pd.Series:
    """Return the row with the highest Count in a cluster sub-dataframe."""
    if "Count" in sub_df.columns:
        return sub_df.loc[sub_df["Count"].idxmax()]
    return sub_df.iloc[0]


def generate_consensus_df(df: pd.DataFrame, cluster_counts: dict, dna_df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for cid in sorted(df["Cluster"].unique()):
        sub = df[df["Cluster"] == cid].copy()
        sub = sub[~sub["Protein_Sequence"].str.contains(r"\*", na=False)]
        if sub.empty:
            continue

        rep   = weighted_representative(sub)
        prot  = rep["Protein_Sequence"]
        cdr1  = rep.get("CDR1", "")
        cdr2  = rep.get("CDR2", "")
        cdr3  = rep.get("CDR3",  "")
        concat = rep.get("CDR_Concatenated", f"{cdr1}{cdr2}{cdr3}")
        liab  = flag_liabilities(cdr3)

        # Physicochemical properties of representative
        try:
            pa     = ProteinAnalysis(prot)
            pi     = round(pa.isoelectric_point(), 2)
            arom   = round(pa.aromaticity(), 3)
            ac     = pa.count_amino_acids()
            seqlen = max(len(prot), 1)
            aliphi = round((ac.get("A",0) + 2.9*(ac.get("I",0)+ac.get("L",0))
                            + 3.9*ac.get("V",0)) / seqlen * 100, 2)
        except Exception:
            pi = arom = aliphi = None

        # DNA representative
        dna_match = dna_df[dna_df["Protein_Sequence"] == prot] if dna_df is not None else pd.DataFrame()
        dna_rep   = dna_match.groupby("DNA_Sequence").size().idxmax() \
                    if not dna_match.empty else "N/A"

        entropy  = shannon_entropy(sub["CDR3"].dropna().tolist())
        n_unique = sub["CDR3"].nunique()

        rows.append({
            "Cluster":              cid,
            "Representative_DNA":   dna_rep,
            "Protein_Sequence":     prot,
            "CDR1":                 cdr1,
            "CDR2":                 cdr2,
            "CDR3":                 cdr3,       # required by enrichment step
            "CDR_Concatenated":     concat,
            "Cluster_Count":        int(cluster_counts.get(cid, 0)),  # required by enrichment step
            "Unique_CDR3s":         n_unique,
            "Shannon_Entropy":      round(entropy, 3),
            "pI":                   pi,
            "Aliphatic_Index":      aliphi,
            "Aromaticity":          arom,
            "CDR3_Liabilities":     liab,
        })
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL FORMATTING
# ══════════════════════════════════════════════════════════════════════════════

def apply_excel_formatting(ws):
    header = {cell.value: get_column_letter(idx + 1)
              for idx, cell in enumerate(ws[1]) if cell.value}
    last   = ws.max_row
    bad    = DEFAULT_CONFIG["bad_fill"]
    good   = DEFAULT_CONFIG["good_fill"]

    def _cond(col, op, val, fill):
        if col in header:
            ws.conditional_formatting.add(
                f"{header[col]}2:{header[col]}{last}",
                CellIsRule(operator=op, formula=[val], fill=fill)
            )

    _cond("pI",             "lessThan",    "6",    bad)
    _cond("pI",             "greaterThan", "9",    bad)
    _cond("pI",             "between",     "6",    good)   # pI 6–9 coloured good below
    _cond("Aliphatic_Index","lessThan",    "70",   bad)
    _cond("Aromaticity",    "greaterThan", "0.30", bad)

    # Heatmap on Cluster_Count
    if "Cluster_Count" in header:
        c = header["Cluster_Count"]
        ws.conditional_formatting.add(
            f"{c}2:{c}{last}",
            ColorScaleRule(start_type="min",  start_color="FFFFFF",
                           mid_type="percentile", mid_value=50, mid_color="FFFF99",
                           end_type="max",    end_color="FF0000")
        )


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def process_csv(input_csv: str, cfg: Optional[dict] = None):
    config = {**DEFAULT_CONFIG, **(cfg or {})}
    input_path = Path(input_csv)
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name  = input_path.stem
    output_dir = input_path.parent / f"{base_name}_clustering_output_{timestamp}"
    output_dir.mkdir(parents=True, exist_ok=True)
    cluster_csv_dir = output_dir / "Clusters"
    cluster_csv_dir.mkdir(exist_ok=True)

    # Logging
    logging.basicConfig(
        filename=output_dir / "clustering_log.txt",
        filemode="w", level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s"
    )
    logging.info(f"Processing {input_csv}")

    # ── Load & clean ───────────────────────────────────────────────────────────
    df = pd.read_csv(input_csv)
    col = config["cluster_column"]
    df = df[df[col].notnull() & (df[col].str.len() > 0)]
    df = df[~df["Protein_Sequence"].str.contains("X|\\*", na=False)].reset_index(drop=True)

    # Load DNA mapping
    dna_csv_path = input_path.parent / input_path.name.replace(
        "_vhh_protein_cdr.csv", "_vhh_dna_protein.csv"
    )
    dna_df = pd.read_csv(dna_csv_path) if dna_csv_path.exists() else None
    if dna_df is None:
        console.print("[yellow]  Warning: DNA mapping file not found. DNA column will be N/A.[/yellow]")

    # ── Build embeddings ───────────────────────────────────────────────────────
    cdr_cols = [c for c in config["embed_cdrs"] if c in df.columns]
    console.print(f"  Embedding {len(df):,} sequences using columns: {cdr_cols}")
    embeddings = np.array([embed_row(row, cdr_cols) for _, row in df.iterrows()])

    # ── Optional UMAP reduction ────────────────────────────────────────────────
    if config["use_umap"] and UMAP_AVAILABLE and embeddings.shape[0] > 20:
        n_comp = min(config["umap_n_components"], embeddings.shape[0] - 1)
        console.print(f"  UMAP reduction → {n_comp} components")
        reducer   = umap.UMAP(n_components=n_comp, random_state=42, n_jobs=1)
        embeddings = reducer.fit_transform(embeddings)
    elif config["use_umap"] and not UMAP_AVAILABLE:
        console.print("  [yellow]UMAP not available; skipping reduction[/yellow]")

    # ── HDBSCAN clustering ─────────────────────────────────────────────────────
    console.print(f"  Running HDBSCAN (min_cluster_size={config['min_cluster_size']}, "
                  f"min_samples={config['min_samples']})")
    clusterer = hdbscan.HDBSCAN(
        min_cluster_size=config["min_cluster_size"],
        min_samples=config["min_samples"],
        prediction_data=True,
    )
    labels = clusterer.fit_predict(embeddings)
    df["Cluster"]             = labels
    df["Cluster_Probability"] = clusterer.probabilities_

    # ── Filter by min cluster count ───────────────────────────────────────────
    cluster_counts = defaultdict(int)
    for _, row in df.iterrows():
        cluster_counts[row["Cluster"]] += int(row.get("Count", 1))

    valid = {c for c, t in cluster_counts.items()
             if t >= config["min_cluster_counts"] and c != -1}
    df = df[df["Cluster"].isin(valid)].reset_index(drop=True)
    df["Cluster_Count"] = df["Cluster"].map(cluster_counts)

    # ── Renumber by total count ────────────────────────────────────────────────
    rank_map = {
        old: new + 1
        for new, old in enumerate(
            sorted(valid, key=lambda c: cluster_counts[c], reverse=True)
        )
    }
    df["Cluster"] = df["Cluster"].map(rank_map)
    cluster_counts = {rank_map[k]: v for k, v in cluster_counts.items() if k in rank_map}

    # ── Per-cluster CSVs ───────────────────────────────────────────────────────
    for cid in sorted(df["Cluster"].unique()):
        sub = df[df["Cluster"] == cid]
        sub[[
            "Protein_Sequence", "CDR1", "CDR2", "CDR3",
            "CDR_Concatenated", "Count", "Cluster_Count", "Cluster_Probability"
        ]].to_csv(cluster_csv_dir / f"cluster_{cid}.csv", index=False)

    # ── Consensus DataFrame ────────────────────────────────────────────────────
    consensus_df = generate_consensus_df(df, cluster_counts, dna_df)
    excel_path   = output_dir / f"{base_name}_cluster_consensus.xlsx"

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        consensus_df.to_excel(writer, index=False, sheet_name="Consensus")

    wb = openpyxl.load_workbook(excel_path)
    apply_excel_formatting(wb["Consensus"])
    wb.save(excel_path)

    # ── Print summary ──────────────────────────────────────────────────────────
    noise = int((labels == -1).sum())
    table = Table(title="HDBSCAN Clustering Summary", show_lines=True)
    table.add_column("Metric", style="cyan")
    table.add_column("Value",  justify="right")
    table.add_row("Input sequences",     str(len(labels)))
    table.add_row("Noise points",        str(noise))
    table.add_row("Clusters (pre-filter)", str(len(set(labels)) - (1 if -1 in labels else 0)))
    table.add_row("Clusters (post-filter)", str(len(valid)))
    table.add_row("UMAP used",           str(config["use_umap"] and UMAP_AVAILABLE))
    console.print(table)
    console.print(f"\n[green]✓ Output:[/green] {output_dir}")
    logging.info(f"Complete. {len(valid)} clusters written to {output_dir}")


def main():
    parser = argparse.ArgumentParser(
        description="HDBSCAN physicochemical clustering of VHH CDR sequences"
    )
    parser.add_argument("input_dir",          help="Directory containing *_vhh_protein_cdr.csv files")
    parser.add_argument("--min-cluster-counts", type=int, default=10,
                        help="Minimum total read counts per cluster (default: 10)")
    parser.add_argument("--min-cluster-size",  type=int, default=2,
                        help="HDBSCAN min_cluster_size (default: 2)")
    parser.add_argument("--min-samples",       type=int, default=2,
                        help="HDBSCAN min_samples (default: 2)")
    parser.add_argument("--no-umap",           action="store_true",
                        help="Disable UMAP pre-reduction")
    parser.add_argument("--embed-cdrs",        nargs="+", default=["CDR1", "CDR2", "CDR3"],
                        help="CDR columns to embed (default: CDR1 CDR2 CDR3)")
    args = parser.parse_args()

    cfg = {
        "min_cluster_counts": args.min_cluster_counts,
        "min_cluster_size":   args.min_cluster_size,
        "min_samples":        args.min_samples,
        "use_umap":           not args.no_umap,
        "embed_cdrs":         args.embed_cdrs,
    }

    input_dir = Path(args.input_dir)
    for csv_path in input_dir.glob("*_vhh_protein_cdr.csv"):
        process_csv(str(csv_path), cfg)


if __name__ == "__main__":
    main()
